# -*- coding: utf-8 -*-
"""
笔记工作台 本地转换服务
======================
把浏览器里做不了的 PDF<->DOCX 转换放到本地 Python 进程：
  * POST /api/pdf2docx   PDF 字节  -> DOCX 字节   （pdf2docx-plus）
  * POST /api/docx2pdf   DOCX 字节 -> PDF 字节    （docx2pdf，走本机 Word）
  * POST /api/excel-edit xlsx 字节 + 单元格数据 -> xlsx 字节（openpyxl 修改）
  * GET  /health         健康检查 / 版本信息

启动：python server/convert_server.py [--port 5198]
默认监听 127.0.0.1:5198，带 CORS，供 http://127.0.0.1:5199 的页面直接调用。
"""
import argparse
import base64
import io
import json
import logging
import os
import re
import subprocess
import sys
import tempfile
import threading
import time
import traceback
import uuid
import warnings
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

warnings.filterwarnings("ignore", category=DeprecationWarning)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("convert-server")

try:
    import pdf2docx_plus  # noqa: F401  (仅用于版本/可用性检测；转换在 convert_worker.py 子进程执行)
except Exception as e:  # noqa: BLE001
    PDF2DOCX_OK_IMPORT = False
else:
    PDF2DOCX_OK_IMPORT = True

# PDF->DOCX 图片保真开关（默认开启，可用环境变量关闭）：
#   NF_RECOVER_MISSING_IMAGES=0   关闭"缺失图片恢复"（修复复用图片对象被丢图，如每页重复 logo）
#   NF_RASTERIZE_VECTORS=0        关闭"矢量图形栅格化"（修复图表/图示等矢量图整块消失）
def _env_flag(name, default=True):
    v = os.environ.get(name)
    if v is None:
        return default
    return v.strip().lower() not in ("0", "false", "no", "off")

_RECOVER_MISSING_IMAGES = _env_flag("NF_RECOVER_MISSING_IMAGES", True)
_RASTERIZE_VECTORS = _env_flag("NF_RASTERIZE_VECTORS", True)

try:
    import docx2pdf  # noqa: F401  (转换时使用)
except Exception as e:  # noqa: BLE001
    DOCX2PDF_OK_IMPORT = False
else:
    DOCX2PDF_OK_IMPORT = True

try:
    from importlib.metadata import version as _pkg_ver
    PDF2DOCX_VER = _pkg_ver("pdf2docx-plus")
except Exception as e:  # noqa: BLE001
    PDF2DOCX_VER = f"missing ({e})"
PDF2DOCX_OK = PDF2DOCX_OK_IMPORT and not PDF2DOCX_VER.startswith("missing")

try:
    from importlib.metadata import version as _pkg_ver2
    DOCX2PDF_VER = _pkg_ver2("docx2pdf")
except Exception as e:  # noqa: BLE001
    DOCX2PDF_VER = f"missing ({e})"
DOCX2PDF_OK = DOCX2PDF_OK_IMPORT and not DOCX2PDF_VER.startswith("missing")

try:
    import openpyxl  # noqa: F401  (xlsx 单元格编辑)
    OPENPYXL_VER = openpyxl.__version__
    OPENPYXL_OK = True
except Exception as e:  # noqa: BLE001
    OPENPYXL_VER = f"missing ({e})"
    OPENPYXL_OK = False

try:
    import win32com.client  # noqa: F401  (docx2pdf 在 Windows 上依赖 pywin32)
    WIN32_OK = True
except Exception:  # noqa: BLE001
    WIN32_OK = False

_MAX_BODY = 256 * 1024 * 1024  # 256MB 上限
_docx2pdf_lock = threading.Lock()
# 转换在独立子进程(convert_worker.py)中执行：进程级隔离，卡死可整体 kill，
# 不会像线程那样留下孤儿线程把服务器进程拖死；并发请求各自独立互不阻塞。
_WORKER_PY = os.path.join(os.path.dirname(os.path.abspath(__file__)), "convert_worker.py")
_PDF2DOCX_TIMEOUT_S = float(os.environ.get("NF_PDF2DOCX_TIMEOUT_S", "180"))  # 单次转换超时上限

# ---- PDF->DOCX 结果缓存 ----
# 同一 PDF（按内容 SHA-256 前缀）转换过一次后，docx 落在 cache/ 目录，
# 下次打开同一文件直接返回缓存的 docx，不再重新转换。
# 前端请求 /api/pdf2docx 时带 ?cache=<hash>；浏览器「设置 → 缓存管理」可查看/多选删除。
_CACHE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "cache")


def _cache_path(key):
    if not key or not re.fullmatch(r"[0-9a-fA-F]{8,64}", key):
        return None
    return os.path.join(_CACHE_DIR, f"{key.lower()}.docx")


def _cache_list():
    """列出缓存 docx：[{key, size, mtime}]（按时间倒序）。"""
    try:
        os.makedirs(_CACHE_DIR, exist_ok=True)
        items = []
        for name in os.listdir(_CACHE_DIR):
            if not name.endswith(".docx"):
                continue
            p = os.path.join(_CACHE_DIR, name)
            try:
                st = os.stat(p)
                items.append({
                    "key": name[:-5],
                    "size": st.st_size,
                    "mtime": st.st_mtime,
                })
            except Exception:  # noqa: BLE001
                continue
        items.sort(key=lambda it: it["mtime"], reverse=True)
        return items
    except Exception:  # noqa: BLE001
        return []


def _cache_delete(keys):
    """删除指定缓存（keys 为 hash 列表），返回删除成功数量。"""
    n = 0
    for k in keys:
        p = _cache_path(k)
        if p is None:
            continue
        try:
            os.remove(p)
            n += 1
        except Exception:  # noqa: BLE001
            pass
    return n

# ---- 转换进度（job 维度）----
# 每个转换请求带一个 job id（?job=xxx），worker/服务端把进度写成 JSON 文件，
# 浏览器在 POST 进行期间并行轮询 GET /api/progress/<job> 拿百分比与阶段文字。
_PROGRESS_DIR = os.path.join(tempfile.gettempdir(), "nf-convert-progress")


def _progress_path(job_id):
    return os.path.join(_PROGRESS_DIR, f"{job_id}.json")


def _write_progress(job_id, payload):
    try:
        os.makedirs(_PROGRESS_DIR, exist_ok=True)
        tmp = _progress_path(job_id) + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        os.replace(tmp, _progress_path(job_id))
    except Exception:  # noqa: BLE001
        pass


def _read_progress(job_id):
    try:
        with open(_progress_path(job_id), "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:  # noqa: BLE001
        return None


def _cleanup_progress(job_id):
    try:
        os.remove(_progress_path(job_id))
    except Exception:  # noqa: BLE001
        pass


def _sweep_stale_progress(max_age_s=7200):
    """清理长时间没被删的进度文件（进程崩溃残留）。"""
    try:
        if not os.path.isdir(_PROGRESS_DIR):
            return
        now = time.time()
        for name in os.listdir(_PROGRESS_DIR):
            p = os.path.join(_PROGRESS_DIR, name)
            try:
                if now - os.path.getmtime(p) > max_age_s:
                    os.remove(p)
            except Exception:  # noqa: BLE001
                pass
    except Exception:  # noqa: BLE001
        pass


class Handler(BaseHTTPRequestHandler):
    server_version = "note-studio-convert/1.0"

    # ---- CORS ----
    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Expose-Headers", "X-NF-Cache-Hit")
        self.send_header("Access-Control-Max-Age", "86400")

    def _send_json(self, code, obj):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self._cors()
        self.end_headers()
        self.wfile.write(body)

    def _send_bytes(self, code, ctype, data):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self._cors()
        self.end_headers()
        self.wfile.write(data)

    def _read_body(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0 or length > _MAX_BODY:
            raise ValueError(f"非法请求体长度: {length}")
        return self.rfile.read(length)

    @staticmethod
    def _is_broken_conn(e):
        # 客户端在中途断开（关标签页/刷新/取消请求），属正常现象，不必当错误上报
        return isinstance(e, (ConnectionAbortedError, ConnectionResetError, BrokenPipeError))

    @staticmethod
    def _job_id_from_path(path):
        """从查询串里取 ?job=xxx（无则 None，调用方再随机生成）。"""
        q = path.split("?", 1)
        if len(q) < 2:
            return None
        for part in q[1].split("&"):
            k, _, v = part.partition("=")
            if k == "job" and v:
                return v[:64]
        return None

    @staticmethod
    def _query_param(path, key):
        """从查询串里取指定参数值（无则 None）。值截断到 64 字符防异常输入。"""
        q = path.split("?", 1)
        if len(q) < 2:
            return None
        for part in q[1].split("&"):
            k, _, v = part.partition("=")
            if k == key and v:
                return v[:64]
        return None

    def log_message(self, fmt, *args):  # 安静一点
        log.info("%s %s", self.address_string(), fmt % args)

    # ---- 路由 ----
    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_GET(self):
        path = self.path.split("?")[0]
        if path == "/health":
            self._send_json(200, {
                "ok": True,
                "pdf2docx": PDF2DOCX_VER,
                "docx2pdf": DOCX2PDF_VER,
                "openpyxl": OPENPYXL_VER,
                "win32": WIN32_OK,
                "word": _word_available(),
            })
        elif path.startswith("/api/progress/"):
            job = path[len("/api/progress/"):]
            data = _read_progress(job)
            if data is None:
                self._send_json(404, {"error": "no progress for job"})
            else:
                self._send_json(200, {"ok": True, **data})
        elif path == "/api/cache":
            # GET /api/cache → 缓存文件列表（设置页显示/多选删除用）
            self._send_json(200, {"ok": True, "files": _cache_list()})
        else:
            self._send_json(404, {"error": "not found"})

    def do_POST(self):
        path = self.path.split("?")[0]
        try:
            if path == "/api/pdf2docx":
                self._handle_pdf2docx()
            elif path == "/api/docx2pdf":
                self._handle_docx2pdf()
            elif path == "/api/excel-edit":
                self._handle_excel_edit()
            elif path == "/api/cache/delete":
                self._handle_cache_delete()
            else:
                self._send_json(404, {"error": "not found"})
        except Exception as e:  # noqa: BLE001
            if self._is_broken_conn(e):
                log.debug("客户端中断连接: %s", e)
                return
            log.error("处理失败: %s\n%s", e, traceback.format_exc())
            try:
                self._send_json(500, {"error": str(e)[:500]})
            except Exception:  # noqa: BLE001
                pass

    def _handle_cache_delete(self):
        """POST /api/cache/delete  body: {"keys": ["hash1", "hash2", ...]}"""
        try:
            payload = json.loads(self._read_body().decode("utf-8"))
        except Exception as e:  # noqa: BLE001
            self._send_json(400, {"error": f"请求体解析失败: {e}"})
            return
        keys = payload.get("keys") or []
        if not isinstance(keys, list):
            self._send_json(400, {"error": "keys 必须是数组"})
            return
        n = _cache_delete([str(k) for k in keys if k])
        log.info("cache/delete: deleted=%d", n)
        self._send_json(200, {"ok": True, "deleted": n})

    def _handle_pdf2docx(self):
        if not PDF2DOCX_OK:
            self._send_json(500, {"error": "pdf2docx-plus 未安装：pip install pdf2docx-plus"})
            return
        job = self._job_id_from_path(self.path) or uuid.uuid4().hex
        cache_key = self._query_param(self.path, "cache")
        data = self._read_body()

        # 缓存命中：同一 PDF（内容 hash 相同）直接返回上次转换的 docx，不重新转换
        cached_path = _cache_path(cache_key) if cache_key else None
        if cached_path and os.path.exists(cached_path) and os.path.getsize(cached_path) > 0:
            try:
                with open(cached_path, "rb") as f:
                    out = f.read()
                log.info("pdf2docx: CACHE HIT %s -> %d bytes", cache_key, len(out))
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
                self.send_header("Content-Length", str(len(out)))
                self.send_header("X-NF-Cache-Hit", "1")
                self._cors()
                self.end_headers()
                self.wfile.write(out)
                return
            except Exception as e:  # noqa: BLE001
                log.warning("缓存读取失败，重新转换: %s", e)

        with tempfile.TemporaryDirectory(prefix="nf-pdf2docx-") as td:
            pdf_path = os.path.join(td, "input.pdf")
            docx_path = os.path.join(td, "output.docx")
            with open(pdf_path, "wb") as f:
                f.write(data)
            timeout_s = _PDF2DOCX_TIMEOUT_S
            worker_err = os.path.join(
                os.path.dirname(_WORKER_PY), "..", "logs", "convert-worker.err.log"
            )
            os.makedirs(os.path.dirname(worker_err), exist_ok=True)
            progress_file = _progress_path(job)
            _write_progress(job, {"percent": 1, "stage": "准备转换…", "done": 0, "total": 0})
            try:
                with open(worker_err, "ab") as errf:
                    proc = subprocess.Popen(
                        [sys.executable, _WORKER_PY, pdf_path, docx_path,
                         "1" if _RECOVER_MISSING_IMAGES else "0",
                         "1" if _RASTERIZE_VECTORS else "0",
                         progress_file],
                        stdout=subprocess.DEVNULL,
                        stderr=errf,
                        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                        cwd=os.path.dirname(_WORKER_PY),
                    )
                    try:
                        rc = proc.wait(timeout=timeout_s)
                    except subprocess.TimeoutExpired:
                        try:
                            proc.kill()
                            proc.wait(timeout=10)
                        except Exception:  # noqa: BLE001
                            pass
                        raise RuntimeError(f"转换超时（>{int(timeout_s)}s），已终止。文件可能过大或结构异常，请重试")
                if rc != 0 or not os.path.exists(docx_path) or os.path.getsize(docx_path) == 0:
                    raise RuntimeError(f"pdf2docx 转换失败（退出码 {rc}）")
                with open(docx_path, "rb") as f:
                    out = f.read()
            finally:
                _cleanup_progress(job)
            # 转换成功 → 写入缓存（下次同文件直接复用；写失败不影响本次返回）
            if cached_path:
                try:
                    os.makedirs(_CACHE_DIR, exist_ok=True)
                    with open(cached_path, "wb") as f:
                        f.write(out)
                    log.info("pdf2docx: cache written %s (%d bytes)", cache_key, len(out))
                except Exception as e:  # noqa: BLE001
                    log.warning("缓存写入失败: %s", e)
        log.info("pdf2docx: %d -> %d bytes", len(data), len(out))
        self._send_bytes(200, "application/vnd.openxmlformats-officedocument.wordprocessingml.document", out)

    def _handle_docx2pdf(self):
        if not DOCX2PDF_OK:
            self._send_json(500, {"error": "docx2pdf 未安装：pip install docx2pdf（需本机安装 Word）"})
            return
        job = self._job_id_from_path(self.path) or uuid.uuid4().hex
        data = self._read_body()
        _write_progress(job, {"percent": 5, "stage": "正在准备…", "done": 0, "total": 0})
        try:
            with _docx2pdf_lock:
                # ThreadingHTTPServer 每个请求在新线程，Word COM 必须先 CoInitialize
                try:
                    import pythoncom
                    pythoncom.CoInitialize()
                except Exception:  # noqa: BLE001
                    pass
                try:
                    _write_progress(job, {"percent": 18, "stage": "正在启动 Word…", "done": 0, "total": 0})
                    with tempfile.TemporaryDirectory(prefix="nf-docx2pdf-") as td:
                        docx_path = os.path.join(td, "input.docx")
                        pdf_path = os.path.join(td, "output.pdf")
                        with open(docx_path, "wb") as f:
                            f.write(data)
                        _write_progress(job, {"percent": 40, "stage": "正在用 Word 生成 PDF…", "done": 0, "total": 0})
                        docx2pdf.convert(docx_path, pdf_path)
                        if not os.path.exists(pdf_path) or os.path.getsize(pdf_path) == 0:
                            raise RuntimeError("docx2pdf 转换失败：未生成 PDF（请确认已安装 Microsoft Word）")
                        _write_progress(job, {"percent": 90, "stage": "正在写出 PDF…", "done": 0, "total": 0})
                        with open(pdf_path, "rb") as f:
                            out = f.read()
                finally:
                    try:
                        import pythoncom
                        pythoncom.CoUninitialize()
                    except Exception:  # noqa: BLE001
                        pass
        finally:
            _cleanup_progress(job)
        log.info("docx2pdf: %d -> %d bytes", len(data), len(out))
        self._send_bytes(200, "application/pdf", out)

    def _handle_excel_edit(self):
        """用 openpyxl 修改 xlsx 并写回（只动被编辑的单元格，保留未编辑内容/格式/多 sheet）。

        请求格式（JSON body）：
          {
            "xlsx": "<base64 原始 xlsx 字节>",
            # 方案 A（兼容旧版）：整体替换第一个工作表
            "rows": [["张三", 95], ...],
            "sheetIndex": 0,
            # 方案 B（编辑操作流，推荐）：逐格/逐行列/工作表级操作
            "wbOps": [
                {"op": "renameSheet", "sheetIndex": 0, "name": "新名字"},
                {"op": "deleteSheet", "sheetIndex": 1},
            ],
            "sheets": [
                {"sheetIndex": 0, "ops": [ {"op": "set", "r": 0, "c": 0, "v": 123, "t": "n"}, ... ]},
                {"sheetIndex": -1, "name": "新建表", "ops": [...]},   # -1 = 新建工作表
            ]
          }
        单元格操作 ops（r/c 均为 0 基）：
          set        {"op":"set","r":0,"c":0,"v":123,"t":"n"}    t: n|b|s|d|e
          insertRow  {"op":"insertRow","at":1,"amount":1}
          deleteRow  {"op":"deleteRow","at":1,"amount":1}
          insertCol  {"op":"insertCol","at":0,"amount":1}
          deleteCol  {"op":"deleteCol","at":0,"amount":1}
          merge      {"op":"merge","r1":0,"c1":0,"r2":1,"c2":1}
          unmerge    {"op":"unmerge","r1":0,"c1":0,"r2":1,"c2":1}
        响应：修改后的 xlsx 字节（二进制）。
        """
        if not OPENPYXL_OK:
            self._send_json(500, {"error": "openpyxl 未安装：pip install openpyxl"})
            return
        try:
            payload = json.loads(self._read_body().decode("utf-8"))
        except Exception as e:  # noqa: BLE001
            self._send_json(400, {"error": f"请求体解析失败: {e}"})
            return
        try:
            data = base64.b64decode(payload.get("xlsx") or "")
        except Exception as e:  # noqa: BLE001
            self._send_json(400, {"error": f"xlsx base64 解码失败: {e}"})
            return
        if not data:
            self._send_json(400, {"error": "缺少 xlsx 字节"})
            return
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
        except Exception as e:  # noqa: BLE001
            self._send_json(400, {"error": f"xlsx 解析失败（仅支持 .xlsx 格式）: {e}"})
            return

        # ---- 方案 A：整体替换（兼容旧版）----
        rows = payload.get("rows")
        if rows is not None:
            if not isinstance(rows, list):
                self._send_json(400, {"error": "rows 必须是数组"})
                return
            try:
                sheet_index = int(payload.get("sheetIndex") or 0)
                ws = wb.worksheets[sheet_index] if sheet_index < len(wb.worksheets) else wb.active
            except Exception:  # noqa: BLE001
                ws = wb.active
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max(ws.max_column, 1)):
                for cell in row:
                    cell.value = None
            for r, row in enumerate(rows, start=1):
                for c, value in enumerate(row, start=1):
                    ws.cell(row=r, column=c, value=value)
            buf = io.BytesIO()
            wb.save(buf)
            out = buf.getvalue()
            log.info("excel-edit: %d -> %d bytes (rows=%d)", len(data), len(out), len(rows))
            self._send_bytes(
                200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", out
            )
            return

        # ---- 方案 B：编辑操作流 ----
        wb_ops = payload.get("wbOps") or []
        sheets = payload.get("sheets") or []
        if not isinstance(wb_ops, list) or not isinstance(sheets, list):
            self._send_json(400, {"error": "wbOps/sheets 必须是数组"})
            return
        try:
            original_count = len(wb.worksheets)
            deleted = set()
            survivors = list(range(original_count))
            for op in wb_ops:
                kind = op.get("op")
                if kind == "renameSheet":
                    idx = int(op.get("sheetIndex") or 0)
                    if idx not in survivors:
                        raise ValueError(f"工作表 {idx} 不存在或已被删除")
                    ws = wb.worksheets[survivors.index(idx)]
                    ws.title = _unique_sheet_title(wb, op.get("name") or "", exclude=idx)
                elif kind == "deleteSheet":
                    idx = int(op.get("sheetIndex") or 0)
                    if idx not in survivors:
                        raise ValueError(f"工作表 {idx} 不存在或已被删除")
                    if len(wb.worksheets) <= 1:
                        raise ValueError("至少保留一个工作表")
                    pos = survivors.index(idx)
                    deleted.add(idx)
                    wb.remove(wb.worksheets[pos])
                    survivors.pop(pos)
                else:
                    raise ValueError(f"未知工作簿操作: {kind}")
            for sheet in sheets:
                idx = sheet.get("sheetIndex")
                ops = sheet.get("ops") or []
                if not isinstance(ops, list):
                    raise ValueError("ops 必须是数组")
                if idx == -1:
                    # 新建工作表：名称由前端保证唯一，服务端再兜底
                    ws = wb.create_sheet(title=_unique_sheet_title(wb, sheet.get("name") or ""))
                else:
                    idx = int(idx or 0)
                    if idx not in survivors:
                        raise ValueError(f"工作表 {idx} 不存在或已被删除")
                    ws = wb.worksheets[survivors.index(idx)]
                for op in ops:
                    _apply_sheet_op(ws, op)
        except Exception as e:  # noqa: BLE001
            self._send_json(400, {"error": str(e)[:500]})
            return
        buf = io.BytesIO()
        wb.save(buf)
        out = buf.getvalue()
        n_ops = sum(len(s.get("ops") or []) for s in sheets)
        log.info("excel-edit: %d -> %d bytes (wbOps=%d sheetOps=%d)",
                 len(data), len(out), len(wb_ops), n_ops)
        self._send_bytes(
            200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", out
        )


def _clean_sheet_title(title):
    """去掉 Excel 工作表名非法字符并截断（Excel 限制：<=31 字符，不能含 []:*?/\\）。"""
    t = (title or "").strip()
    for ch in r'[]:*?/\\':
        t = t.replace(ch, "")
    return t[:31].strip() or "Sheet"


def _unique_sheet_title(wb, title, exclude=None):
    """在现有工作表名中取唯一名称（同名自动追加数字）。exclude 为原索引，跳过自身。"""
    base = _clean_sheet_title(title)
    names = {ws.title for i, ws in enumerate(wb.worksheets) if i != exclude}
    if base not in names:
        return base
    i = 1
    while "{0}{1}".format(base, i) in names:
        i += 1
    return "{0}{1}".format(base, i)


def _apply_sheet_op(ws, op):
    """把一条单元格/行列操作应用到 openpyxl 工作表（r/c 均为 0 基）。"""
    kind = op.get("op")

    def target(r, c):
        return ws.cell(row=r + 1, column=c + 1)

    if kind == "set":
        r = int(op.get("r") or 0)
        c = int(op.get("c") or 0)
        v = op.get("v")
        t = op.get("t") or "s"
        if t == "e" or v is None:
            target(r, c).value = None
        elif t == "n":
            try:
                fv = float(v)
                target(r, c).value = int(fv) if fv.is_integer() else fv
            except Exception:  # noqa: BLE001
                target(r, c).value = str(v)
        elif t == "b":
            target(r, c).value = bool(v)
        elif t == "d":
            try:
                from datetime import datetime
                target(r, c).value = datetime.fromisoformat(str(v))
            except Exception:  # noqa: BLE001
                target(r, c).value = str(v)
        else:
            # 字符串；以 '=' 开头会被 openpyxl 当作公式
            target(r, c).value = str(v)
    elif kind == "insertRow":
        ws.insert_rows(int(op.get("at") or 0) + 1, max(1, int(op.get("amount") or 1)))
    elif kind == "deleteRow":
        ws.delete_rows(int(op.get("at") or 0) + 1, max(1, int(op.get("amount") or 1)))
    elif kind == "insertCol":
        ws.insert_cols(int(op.get("at") or 0) + 1, max(1, int(op.get("amount") or 1)))
    elif kind == "deleteCol":
        ws.delete_cols(int(op.get("at") or 0) + 1, max(1, int(op.get("amount") or 1)))
    elif kind == "merge":
        ws.merge_cells(
            start_row=int(op["r1"]) + 1, start_column=int(op["c1"]) + 1,
            end_row=int(op["r2"]) + 1, end_column=int(op["c2"]) + 1,
        )
    elif kind == "unmerge":
        ws.unmerge_cells(
            start_row=int(op["r1"]) + 1, start_column=int(op["c1"]) + 1,
            end_row=int(op["r2"]) + 1, end_column=int(op["c2"]) + 1,
        )
    else:
        raise ValueError("未知单元格操作: {0}".format(kind))


def _word_available():
    try:
        import winreg
        with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Word.Application\CurVer"):
            return True
    except Exception:  # noqa: BLE001
        return False


def main():
    ap = argparse.ArgumentParser(description="PDF<->DOCX 本地转换服务")
    ap.add_argument("--port", type=int, default=5198)
    ap.add_argument("--host", default="127.0.0.1")
    args = ap.parse_args()
    _sweep_stale_progress()
    log.info("pdf2docx-plus=%s docx2pdf=%s openpyxl=%s win32=%s word=%s",
             PDF2DOCX_VER, DOCX2PDF_VER, OPENPYXL_VER, WIN32_OK, _word_available())
    srv = ThreadingHTTPServer((args.host, args.port), Handler)
    log.info("转换服务已启动: http://%s:%d  (Ctrl+C 停止)", args.host, args.port)
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        srv.shutdown()


if __name__ == "__main__":
    main()
